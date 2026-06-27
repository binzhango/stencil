import importlib
import json
import re
import zipfile
from collections.abc import Mapping
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from pathlib import Path
from subprocess import CompletedProcess, TimeoutExpired
from xml.etree import ElementTree
from xml.sax.saxutils import escape

import pytest

from stencil import UnsupportedFormatError, render
from stencil.errors import RenderError
from stencil.pdf import LibreOfficePdfConverter, PdfConversionError

render_module = importlib.import_module("stencil.render")
WORD_TEXT = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"
PPTX_TEXT = "{http://schemas.openxmlformats.org/drawingml/2006/main}t"
PRESENTATION_SLIDE_ID = "{http://schemas.openxmlformats.org/presentationml/2006/main}sldId"
RELATIONSHIP = "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
REL_ID = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"


def _rendered_docx_text(document: bytes) -> str:
    with zipfile.ZipFile(BytesIO(document)) as archive:
        parts: list[str] = []
        for name in sorted(archive.namelist()):
            if not name.startswith("word/") or not name.endswith(".xml"):
                continue
            root = ElementTree.fromstring(archive.read(name))
            text = " ".join(
                node.text for node in root.iter(WORD_TEXT) if node.text is not None
            )
            if text.strip():
                parts.append(text)
    return " ".join(parts)


def _load_rendered_workbook(document: bytes):
    from openpyxl import load_workbook

    return load_workbook(BytesIO(document), data_only=False)


def _write_minimal_pptx(path: Path, slide_texts: list[str]) -> None:
    content_type_overrides = "\n".join(
        (
            f'<Override PartName="/ppt/slides/slide{index}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
        )
        for index in range(1, len(slide_texts) + 1)
    )
    presentation_slide_ids = "\n".join(
        f'<p:sldId id="{255 + index}" r:id="rId{index}"/>'
        for index in range(1, len(slide_texts) + 1)
    )
    presentation_relationships = "\n".join(
        (
            f'<Relationship Id="rId{index}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" '
            f'Target="slides/slide{index}.xml"/>'
        )
        for index in range(1, len(slide_texts) + 1)
    )

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            (
                '<?xml version="1.0" encoding="UTF-8"?>'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Default Extension="rels" '
                'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                '<Default Extension="xml" ContentType="application/xml"/>'
                '<Override PartName="/ppt/presentation.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>'
                f"{content_type_overrides}"
                "</Types>"
            ),
        )
        archive.writestr(
            "_rels/.rels",
            (
                '<?xml version="1.0" encoding="UTF-8"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                'relationships/officeDocument" '
                'Target="ppt/presentation.xml"/>'
                "</Relationships>"
            ),
        )
        archive.writestr(
            "ppt/presentation.xml",
            (
                '<?xml version="1.0" encoding="UTF-8"?>'
                '<p:presentation xmlns:p="http://schemas.openxmlformats.org/'
                'presentationml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                f"<p:sldIdLst>{presentation_slide_ids}</p:sldIdLst>"
                "</p:presentation>"
            ),
        )
        archive.writestr(
            "ppt/_rels/presentation.xml.rels",
            (
                '<?xml version="1.0" encoding="UTF-8"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                f"{presentation_relationships}"
                "</Relationships>"
            ),
        )
        for index, text in enumerate(slide_texts, start=1):
            archive.writestr(
                f"ppt/slides/slide{index}.xml",
                (
                    '<?xml version="1.0" encoding="UTF-8"?>'
                    '<p:sld xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" '
                    'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
                    "<p:cSld><p:spTree><p:sp><p:txBody><a:bodyPr/><a:lstStyle/>"
                    f"<a:p><a:r><a:t>{escape(text)}</a:t></a:r></a:p>"
                    "</p:txBody></p:sp></p:spTree></p:cSld>"
                    "</p:sld>"
                ),
            )


def _rendered_pptx_slide_texts(document: bytes) -> list[str]:
    with zipfile.ZipFile(BytesIO(document)) as archive:
        presentation = ElementTree.fromstring(archive.read("ppt/presentation.xml"))
        rels = ElementTree.fromstring(archive.read("ppt/_rels/presentation.xml.rels"))
        targets = {
            relationship.attrib["Id"]: relationship.attrib["Target"]
            for relationship in rels.iter(RELATIONSHIP)
        }

        texts: list[str] = []
        for slide_id in presentation.iter(PRESENTATION_SLIDE_ID):
            slide_path = f"ppt/{targets[slide_id.attrib[REL_ID]]}"
            root = ElementTree.fromstring(archive.read(slide_path))
            texts.append(
                " ".join(node.text for node in root.iter(PPTX_TEXT) if node.text)
            )
        return texts


def test_render_rejects_unsupported_template_suffix() -> None:
    with pytest.raises(UnsupportedFormatError):
        render("template.odp", {})


def test_render_rejects_unsupported_output_format() -> None:
    with pytest.raises(UnsupportedFormatError):
        render("template.docx", {}, output_format="xlsx")


def test_xlsx_simple_substitutions_preserve_types_and_styles(tmp_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice"
    sheet["A1"] = "Invoice {{ invoice.id }}"
    sheet["A1"].font = Font(bold=True)
    sheet["B2"] = "{{ invoice.total }}"
    sheet["B2"].number_format = "$#,##0.00"
    sheet["A3"] = "{{ invoice.note }}"
    workbook.create_sheet("Summary")["A1"] = "{{ customer.name }}"
    workbook.save(template)

    rendered = render(
        template,
        {
            "invoice": {"id": "INV-1001", "total": 1800, "note": "A&B <ok>"},
            "customer": {"name": "Acme Studio"},
        },
    )

    rendered_workbook = _load_rendered_workbook(rendered)
    sheet = rendered_workbook["Invoice"]
    assert sheet["A1"].value == "Invoice INV-1001"
    assert sheet["A1"].font.bold is True
    assert sheet["B2"].value == 1800
    assert sheet["B2"].number_format == "$#,##0.00"
    assert sheet["A3"].value == "A&B <ok>"
    assert rendered_workbook["Summary"]["A1"].value == "Acme Studio"


def test_xlsx_row_loop_clones_rows_and_translates_formulas(tmp_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Description"
    sheet["B1"] = "Qty"
    sheet["C1"] = "Unit"
    sheet["D1"] = "Amount"
    sheet["A2"] = "{% for item in line_items %}"
    sheet["A3"] = "{{ item.description }}"
    sheet["B3"] = "{{ item.quantity }}"
    sheet["C3"] = "{{ item.unit_price }}"
    sheet["D3"] = "=B3*C3"
    sheet["A3"].fill = PatternFill("solid", fgColor="FFFF00")
    sheet["A4"] = "{% endfor %}"
    sheet["C5"] = "Total"
    sheet["D5"] = "=SUM(D3:D3)"
    workbook.save(template)

    rendered = render(
        template,
        {
            "line_items": [
                {"description": "Implementation", "quantity": 1, "unit_price": 1200},
                {"description": "QA", "quantity": 2, "unit_price": 300},
            ]
        },
    )

    rendered_workbook = _load_rendered_workbook(rendered)
    sheet = rendered_workbook.active
    assert sheet["A2"].value == "Implementation"
    assert sheet["B2"].value == 1
    assert sheet["C2"].value == 1200
    assert sheet["D2"].value == "=B2*C2"
    assert sheet["A3"].value == "QA"
    assert sheet["B3"].value == 2
    assert sheet["C3"].value == 300
    assert sheet["D3"].value == "=B3*C3"
    assert sheet["A2"].fill.fgColor.rgb == "00FFFF00"
    assert sheet["C4"].value == "Total"
    assert sheet["D4"].value == "=SUM(D3:D3)"


def test_pptx_simple_substitutions_preserve_slide_order(tmp_path: Path) -> None:
    template = tmp_path / "template.pptx"
    _write_minimal_pptx(
        template,
        [
            "Status for {{ customer.name }}",
            "{% if project.on_track %}On track{% else %}At risk{% endif %}",
        ],
    )

    rendered = render(
        template,
        {"customer": {"name": "Acme Studio"}, "project": {"on_track": True}},
    )

    assert _rendered_pptx_slide_texts(rendered) == [
        "Status for Acme Studio",
        "On track",
    ]


def test_pptx_slide_loop_clones_slides_and_rewrites_relationships(tmp_path: Path) -> None:
    template = tmp_path / "template.pptx"
    _write_minimal_pptx(
        template,
        [
            "Executive summary",
            "{% for milestone in milestones %}",
            "{{ milestone.name }}: {{ milestone.status }}",
            "{% endfor %}",
            "Thank you",
        ],
    )

    rendered = render(
        template,
        {
            "milestones": [
                {"name": "Design", "status": "Done"},
                {"name": "Build", "status": "Active"},
            ]
        },
    )

    assert _rendered_pptx_slide_texts(rendered) == [
        "Executive summary",
        "Design: Done",
        "Build: Active",
        "Thank you",
    ]

    with zipfile.ZipFile(BytesIO(rendered)) as archive:
        presentation = archive.read("ppt/presentation.xml").decode("utf-8")
        relationships = archive.read("ppt/_rels/presentation.xml.rels").decode("utf-8")
        assert len(re.findall(r"<p:sldId\b", presentation)) == 4
        assert "slides/slide6.xml" in relationships
        assert "slides/slide7.xml" in relationships


def test_render_converts_pptx_to_pdf(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(render_module, "render_pptx", lambda path, data: b"pptx")
    monkeypatch.setattr(render_module, "convert_pptx_to_pdf", lambda source: b"%PDF slides")

    assert render("template.pptx", {}, output_format="pdf") == b"%PDF slides"


def test_render_converts_xlsx_to_pdf(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(render_module, "render_xlsx", lambda path, data: b"xlsx")
    monkeypatch.setattr(render_module, "convert_xlsx_to_pdf", lambda source: b"%PDF spreadsheet")

    assert render("template.xlsx", {}, output_format="pdf") == b"%PDF spreadsheet"


def test_render_rejects_missing_docx_template() -> None:
    with pytest.raises(RenderError, match="Template file does not exist"):
        render(Path("missing.docx"), {})


def test_invoice_example_matches_expected_output_text() -> None:
    payload = json.loads(Path("examples/data/invoice.json").read_text(encoding="utf-8"))

    rendered = render("examples/templates/invoice.docx", payload)

    text = _rendered_docx_text(rendered)
    assert "{{" not in text
    assert "{%" not in text
    assert "Invoice INV-1001" in text
    assert "Customer: Acme Studio (hello@example.com)" in text
    assert "Template implementation x 1: $1,200.00" in text
    assert "DOCX rendering QA x 2: $300.00" in text
    assert "Total: $1,800.00" in text
    assert "Note: Thanks for building with Stencil." in text


def test_styled_status_report_example_matches_expected_output_text() -> None:
    payload = json.loads(Path("examples/data/status-report.json").read_text(encoding="utf-8"))

    rendered = render("examples/templates/styled-status-report.docx", payload)

    text = _rendered_docx_text(rendered)
    assert "{{" not in text
    assert "{%" not in text
    assert "STENCIL STATUS" in text
    assert "Project Status Report" in text
    assert "Stencil DOCX MVP" in text
    assert "The template defines the layout and styles." in text
    assert "Owner Bin Zhang" in text
    assert "Status On track" in text
    assert "Create styled template Design 2026-06-27" in text
    assert "Render with JSON data Engineering 2026-06-28" in text
    assert "Review output formatting Operations 2026-06-29" in text
    assert "Risk note: Do not put style instructions in JSON" in text


def test_pptx_status_example_matches_expected_output_text() -> None:
    payload = json.loads(Path("examples/data/pptx-status.json").read_text(encoding="utf-8"))

    rendered = render("examples/templates/pptx-status.pptx", payload)

    assert _rendered_pptx_slide_texts(rendered) == [
        "STENCIL EXAMPLE Stencil PPTX Engine Owner: Platform Tools "
        "Status: Ready for internal review "
        "This is a template deck; render it to create a presentation.",
        "MILESTONE Text placeholders Complete Slide text is rendered with Jinja data.",
        "MILESTONE Repeated slides Complete Marker slides expand one detail slide per milestone.",
        "MILESTONE PDF export Available "
        "LibreOffice conversion uses the same PDF worker as DOCX and XLSX.",
        "NEXT STEP Open the rendered deck Confirm the milestone detail slide was repeated once "
        "for each item in the JSON data.",
    ]


def test_render_converts_docx_to_pdf(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(render_module, "render_docx", lambda path, data: b"docx")
    monkeypatch.setattr(render_module, "convert_docx_to_pdf", lambda source: b"%PDF rendered")

    assert render("template.docx", {}, output_format="pdf") == b"%PDF rendered"


def test_render_pdf_dispatch_is_safe_for_concurrent_calls(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fake_render_docx(path: Path, data: Mapping[str, object]) -> bytes:
        return f"docx-{data['value']}".encode()

    def fake_convert_docx_to_pdf(source: bytes) -> bytes:
        return b"%PDF " + source

    monkeypatch.setattr(render_module, "render_docx", fake_render_docx)
    monkeypatch.setattr(render_module, "convert_docx_to_pdf", fake_convert_docx_to_pdf)

    with ThreadPoolExecutor(max_workers=4) as executor:
        results = list(
            executor.map(
                lambda value: render("template.docx", {"value": value}, output_format="pdf"),
                range(12),
            )
        )

    assert sorted(results) == sorted(f"%PDF docx-{value}".encode() for value in range(12))


def test_libreoffice_converter_writes_pdf_and_cleans_temp_workspace(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("stencil.pdf.shutil.which", lambda name: "/usr/bin/soffice")

    def fake_run(command: list[str], **kwargs: object) -> CompletedProcess[list[str]]:
        outdir = Path(command[command.index("--outdir") + 1])
        outdir.joinpath("source.pdf").write_bytes(b"%PDF success")
        return CompletedProcess(command, 0, stdout=b"converted", stderr=b"")

    monkeypatch.setattr("stencil.pdf.subprocess.run", fake_run)

    converter = LibreOfficePdfConverter(timeout_seconds=1, temp_root=tmp_path)

    assert converter.convert(b"docx", source_suffix=".docx") == b"%PDF success"
    assert list(tmp_path.iterdir()) == []


def test_libreoffice_converter_retries_transient_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("stencil.pdf.shutil.which", lambda name: "/usr/bin/soffice")
    attempts = 0

    def fake_run(command: list[str], **kwargs: object) -> CompletedProcess[list[str]]:
        nonlocal attempts
        attempts += 1
        if attempts == 1:
            return CompletedProcess(command, 1, stdout=b"", stderr=b"temporary failure")

        outdir = Path(command[command.index("--outdir") + 1])
        outdir.joinpath("source.pdf").write_bytes(b"%PDF after retry")
        return CompletedProcess(command, 0, stdout=b"converted", stderr=b"")

    monkeypatch.setattr("stencil.pdf.subprocess.run", fake_run)

    converter = LibreOfficePdfConverter(timeout_seconds=1, retries=1, temp_root=tmp_path)

    assert converter.convert(b"docx", source_suffix=".docx") == b"%PDF after retry"
    assert attempts == 2
    assert list(tmp_path.iterdir()) == []


def test_libreoffice_converter_retries_after_timeout(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("stencil.pdf.shutil.which", lambda name: "/usr/bin/soffice")
    attempts = 0

    def fake_run(command: list[str], **kwargs: object) -> CompletedProcess[list[str]]:
        nonlocal attempts
        attempts += 1
        if attempts == 1:
            raise TimeoutExpired(cmd=command, timeout=1)

        outdir = Path(command[command.index("--outdir") + 1])
        outdir.joinpath("source.pdf").write_bytes(b"%PDF after timeout")
        return CompletedProcess(command, 0, stdout=b"converted", stderr=b"")

    monkeypatch.setattr("stencil.pdf.subprocess.run", fake_run)

    converter = LibreOfficePdfConverter(timeout_seconds=1, retries=1, temp_root=tmp_path)

    assert converter.convert(b"docx", source_suffix=".docx") == b"%PDF after timeout"
    assert attempts == 2
    assert list(tmp_path.iterdir()) == []


def test_libreoffice_converter_reports_missing_executable(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("stencil.pdf.shutil.which", lambda name: None)

    converter = LibreOfficePdfConverter()

    with pytest.raises(PdfConversionError, match="requires LibreOffice"):
        converter.convert(b"docx", source_suffix=".docx")
