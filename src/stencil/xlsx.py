"""XLSX rendering engine."""

from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

from jinja2 import Environment, StrictUndefined, TemplateError

from .errors import RenderError

CELL_EXPRESSION_RE = re.compile(r"^\s*\{\{\s*(?P<expression>.*?)\s*\}\}\s*$", re.DOTALL)
FOR_TAG_RE = re.compile(
    r"^\s*\{%\s*for\s+(?P<variable>[A-Za-z_][A-Za-z0-9_]*)\s+in\s+"
    r"(?P<expression>.*?)\s*%\}\s*$",
    re.DOTALL,
)
END_FOR_TAG_RE = re.compile(r"^\s*\{%\s*endfor\s*%\}\s*$", re.DOTALL)


def render_xlsx(template_path: str | Path, data: Mapping[str, Any]) -> bytes:
    """Render an XLSX template with Jinja-compatible data."""

    path = Path(template_path)

    if not path.exists():
        raise RenderError("Template file does not exist", template_path=path, stage="load")
    if not path.is_file():
        raise RenderError("Template path is not a file", template_path=path, stage="load")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RenderError(
            "XLSX rendering requires the 'openpyxl' dependency; run 'uv sync --dev'",
            template_path=path,
            stage="import",
        ) from exc

    try:
        workbook = load_workbook(path)
    except Exception as exc:
        raise RenderError("Failed to load XLSX template", template_path=path, stage="load") from exc

    environment = Environment(undefined=StrictUndefined, autoescape=False)
    context = dict(data)

    try:
        for worksheet in workbook.worksheets:
            _render_worksheet(worksheet, context, environment)
    except TemplateError as exc:
        raise RenderError(
            f"Failed to render XLSX template: {exc}",
            template_path=path,
            stage="render",
        ) from exc
    except Exception as exc:
        raise RenderError(
            "Failed to render XLSX template", template_path=path, stage="render"
        ) from exc

    buffer = BytesIO()
    try:
        workbook.save(buffer)
    except Exception as exc:
        raise RenderError(
            "Failed to serialize rendered XLSX", template_path=path, stage="save"
        ) from exc

    return buffer.getvalue()


def _render_worksheet(worksheet: Any, context: Mapping[str, Any], environment: Environment) -> None:
    for block in reversed(_find_loop_blocks(worksheet)):
        _render_loop_block(worksheet, block, context, environment)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = _render_cell_value(
                cell.value,
                context=context,
                environment=environment,
                formula_origin=cell.coordinate,
                formula_target=cell.coordinate,
            )


def _find_loop_blocks(worksheet: Any) -> list[tuple[int, int, str, str]]:
    stack: list[tuple[int, str, str]] = []
    blocks: list[tuple[int, int, str, str]] = []

    for row_index in range(1, worksheet.max_row + 1):
        row_tags = [
            value
            for value in (_single_text_cell_value(cell) for cell in worksheet[row_index])
            if value is not None
        ]
        for value in row_tags:
            for_match = FOR_TAG_RE.match(value)
            if for_match:
                stack.append(
                    (
                        row_index,
                        for_match.group("variable"),
                        for_match.group("expression").strip(),
                    )
                )
                continue

            if END_FOR_TAG_RE.match(value):
                if not stack:
                    raise ValueError(
                        f"Found endfor without matching for in worksheet {worksheet.title}"
                    )
                start_row, variable, expression = stack.pop()
                if start_row >= row_index - 1:
                    raise ValueError(
                        "XLSX row loop in worksheet "
                        f"{worksheet.title} must contain at least one body row"
                    )
                blocks.append((start_row, row_index, variable, expression))

    if stack:
        start_row, _, _ = stack[-1]
        raise ValueError(
            f"Found for without matching endfor at row {start_row} in {worksheet.title}"
        )

    return blocks


def _single_text_cell_value(cell: Any) -> str | None:
    value = cell.value
    if isinstance(value, str):
        stripped = value.strip()
        if FOR_TAG_RE.match(stripped) or END_FOR_TAG_RE.match(stripped):
            return stripped
    return None


def _render_loop_block(
    worksheet: Any,
    block: tuple[int, int, str, str],
    context: Mapping[str, Any],
    environment: Environment,
) -> None:
    start_row, end_row, variable, expression = block
    body_templates = [
        _capture_row(worksheet, row_index) for row_index in range(start_row + 1, end_row)
    ]
    row_heights = {
        row_index: worksheet.row_dimensions[row_index].height
        for row_index in range(start_row + 1, end_row)
    }
    items = _evaluate_iterable(expression, context, environment)
    body_row_count = len(body_templates)
    output_row_count = body_row_count * len(items)

    worksheet.delete_rows(start_row, end_row - start_row + 1)
    if output_row_count:
        worksheet.insert_rows(start_row, output_row_count)

    for item_index, item in enumerate(items):
        loop_context = dict(context)
        loop_context[variable] = item
        for template_index, row_template in enumerate(body_templates):
            target_row = start_row + (item_index * body_row_count) + template_index
            source_row = start_row + 1 + template_index
            if row_heights[source_row] is not None:
                worksheet.row_dimensions[target_row].height = row_heights[source_row]
            for source_cell in row_template:
                target = worksheet.cell(row=target_row, column=source_cell.column)
                _copy_cell(source_cell, target)
                target.value = _render_cell_value(
                    source_cell.value,
                    context=loop_context,
                    environment=environment,
                    formula_origin=source_cell.coordinate,
                    formula_target=target.coordinate,
                )


def _capture_row(worksheet: Any, row_index: int) -> list[Any]:
    return [copy(cell) for cell in worksheet[row_index]]


def _copy_cell(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)
    if source.hyperlink:
        target._hyperlink = copy(source.hyperlink)
    if source.comment:
        target.comment = copy(source.comment)


def _render_cell_value(
    value: Any,
    *,
    context: Mapping[str, Any],
    environment: Environment,
    formula_origin: str,
    formula_target: str,
) -> Any:
    if not isinstance(value, str):
        return value

    if "{{" in value or "{%" in value:
        expression_match = CELL_EXPRESSION_RE.match(value)
        if expression_match:
            expression = expression_match.group("expression")
            return environment.compile_expression(expression)(**context)
        return environment.from_string(value).render(dict(context))

    if value.startswith("=") and formula_origin != formula_target:
        try:
            from openpyxl.formula.translate import Translator

            return Translator(value, origin=formula_origin).translate_formula(formula_target)
        except Exception:
            return value

    return value


def _evaluate_iterable(
    expression: str, context: Mapping[str, Any], environment: Environment
) -> Sequence[Any]:
    value = environment.compile_expression(expression)(**context)
    if isinstance(value, (str, bytes)) or not isinstance(value, Sequence):
        raise ValueError(f"XLSX row loop expression {expression!r} did not evaluate to a list")
    return value
